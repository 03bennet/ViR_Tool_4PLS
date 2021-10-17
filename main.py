import time
import os
import ntpath
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Border, Alignment, Side
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter import filedialog
import PySimpleGUI as sg


# list of table headers in which data from xml is pulled into

s_p_x, s_p_y, s_p_z, back_str, ahead_st, w_case, h_a_grd, radial_dist  = [], [], [], [], [], [], [], []

# list of file paths to be fed into xml conversion script:

GIAS_path, GIBO_path, GIMO_path, FIMO_path = [], [], [], []

# report naming variable that changes depending on which report is running

report_naming = []

# workbook and sheet variables ready for copying to master template

GIAS_report = []
GIAS_report_sheet = []

GIBO_report = []
GIBO_report_sheet = []

GIMO_report = []
GIMO_report_sheet = []

FIMO_report = []
FIMO_report_sheet = []


# button commands

def GIAS_import():
    global GIAS_path
    GIAS_path = filedialog.askopenfilename()
    if ".xml" in GIAS_path:
        feedback_text_1.configure(text = ntpath.basename(GIAS_path), bg = "OliveDrab4")
    else:
        errormsg(msg = "File is not an XML")


def GIBO_import():
    global GIBO_path
    GIBO_path = filedialog.askopenfilename()
    if ".xml" in GIBO_path:
        feedback_text_2.configure(text = ntpath.basename(GIBO_path), bg = "OliveDrab4")
    else:
        errormsg(msg = "File is not an XML")


def GIMO_import():
    global GIMO_path
    GIMO_path = filedialog.askopenfilename()
    if ".xml" in GIMO_path:
        feedback_text_3.configure(text = ntpath.basename(GIMO_path), bg = "OliveDrab4")
    else:
        errormsg(msg = "File is not an XML")


def FIMO_import():
    global FIMO_path
    FIMO_path = filedialog.askopenfilename()
    if ".xml" in FIMO_path:
        feedback_text_4.configure(text = ntpath.basename(FIMO_path), bg = "OliveDrab4")
    else:
        errormsg(msg = "File is not an XML")


def xml_conversion_script(xml_path):

	# removal of quotation marks if filepath contains them

	if '"' in xml_path:
		xml_path = xml_path.replace('"','')

	# determines whether xml has infringements
	with open(xml_path, "r") as xml:
	    mytree = ET.parse(xml) 
	    myroot = mytree.getroot()
	        
	    if myroot.find('table'):
	        has_data = True

	    else:
	        has_data = False

	# if the xml does have infringement data, the script is run
	# xml table is found and output as an xlsx for easy parsing
	if has_data == True:

		feedback_text_5.configure(text = f"Processing")
		try:
			xml = pd.read_xml(xml_path, "/root/table/vegetation_analysis_report","", True, False ) 
			xml.to_excel("dataframe.xlsx", index = False)
			workbook = pd.read_excel("dataframe.xlsx")
		except:
			errormsg(msg = "Error reading or writing files")
			time.sleep(20)
			exit()

		# xlsx columns from workbook are assigned to the list of table headers
		# if statments are to determine whether report is grow in or fall in
		s_p_x = (workbook['survey_point_x'])
		s_p_y = (workbook['survey_point_y'])
		s_p_z = (workbook['survey_point_z_h'])
		if xml_path == GIAS_path or xml_path == GIBO_path or xml_path == GIMO_path:
			back_str = (workbook['grow_in_closest_wire_back_str'])
		if xml_path == FIMO_path :
			back_str = (workbook['falling_tree_closest_wire_back_str'])
		if xml_path == GIAS_path or xml_path == GIBO_path or xml_path == GIMO_path:
			ahead_st = (workbook['grow_in_closest_wire_ahead_str'])
		if xml_path == FIMO_path :
			ahead_st = (workbook['falling_tree_closest_wire_ahead_str'])
		if xml_path == GIAS_path or xml_path == GIBO_path or xml_path == GIMO_path:
			w_case = (workbook['grow_in_closest_wire_weather_case'])
		if xml_path == FIMO_path :
			w_case = (workbook['falling_tree_closest_wire_weather_case'])
		h_a_grd = (workbook['survey_point_height_above_ground'])
		if xml_path == GIAS_path or xml_path == GIBO_path or xml_path == GIMO_path:
			radial_dist = (workbook['grow_in_closest_wire_vert_or_radial_margin'])
		if xml_path == FIMO_path :
			radial_dist = (workbook['falling_tree_closest_wire_radial_margin'])

		# list is then assigned to dataframe headers
		df = pd.DataFrame.from_dict({'survey_point_x': s_p_x,'survey_point_y': s_p_y, 
		"survey_point_z": s_p_z, "back_structure": back_str, "ahead_structure": ahead_st,
		"weather": w_case, "height_above_ground": h_a_grd, "radial_distance": radial_dist})

		# df is output as an xlsx, basename of filepath is used for naming convention
		df.to_excel(report_naming + '.xlsx', header=True, index=False)


# function that copies the cell range of each report and copies it to the relevant sheet
# on the the master template. Outputs a new file.
def xlsx_copy(reportsheet, destsheet):

	reportsheet.delete_rows(1)

	mr = reportsheet.max_row
	mc = reportsheet.max_column

	for i in range(1, mr + 1):
		for j in range (1, mc + 1):
			c = reportsheet.cell(row = i, column = j)
			destsheet.cell(row = i+9, column = j).value = c.value
			
	template.save('Final_Report.xlsx')


# function checks if final spreadsheet already exists. If it does it deletes before proceeding.
# function then feeds all reports into xlsx_copy function
def xlsx_collate():

	if os.path.isfile('Final_Report.xlsx'):
		os.remove('Final_Report.xlsx')

	if os.path.isfile('Grow-In_As-Surveyed.xlsx'):
		xlsx_copy(GIAS_report_sheet, GIAS_sheet)
	if os.path.isfile('Grow-In_Blow-Out.xlsx'):
		xlsx_copy(GIBO_report_sheet, GIBO_sheet)
	if os.path.isfile('Grow-In_Max-Op.xlsx'):
		xlsx_copy(GIMO_report_sheet, GIMO_sheet)
	if os.path.isfile('Fall-In-Max-Op.xlsx'):
		xlsx_copy(FIMO_report_sheet, FIMO_sheet)
	
	#formatting function is called here if there is a final report xlsx
	if os.path.isfile('Final_Report.xlsx'):
		xlsx_format('Grow-In (As-Surveyed MVCD)')
		xlsx_format('Blow-Out (NESC Blowout WSZ)')
		xlsx_format('Grow-In(CircuitThermal STE WSZ)')
		xlsx_format('Fall-In(CircuitThermal STE WSZ)')
	else:
		errormsg(msg = "\nError! Could not find 'Final_Report.xlsx'")

#function centers text and creates borders

def xlsx_format(worksheet):

    #defines the border style and then asigns it to each border variable
    all_sides = Side(border_style='thin')
    border = Border(left=all_sides, right=all_sides)
    b_border = Border(left=all_sides, right=all_sides, bottom=all_sides)

    #loads relevant page on spreadsheet
    xlsx_file = 'Final_Report.xlsx'

    wb = load_workbook(filename = xlsx_file)

    ws = wb[worksheet]

    #iterates through rows and adds borders
    for row in ws[10:ws.max_row]:
        for i in row:
            i.alignment = Alignment(horizontal="center")
            i.border = border
    #finds max row of spreadsheet and concatonates number to column number range
    max_row =(len(ws['A']))

    row_range = 'A'+str(max_row), 'B'+str(max_row), 'C'+str(max_row), 'D'+str(max_row), 'E'+str(max_row), 'F'+str(max_row), 'G'+str(max_row), 'H'+str(max_row)

    #applies bottom border
    ws[row_range[0]].border = b_border
    ws[row_range[1]].border = b_border
    ws[row_range[2]].border = b_border
    ws[row_range[3]].border = b_border
    ws[row_range[4]].border = b_border
    ws[row_range[5]].border = b_border
    ws[row_range[6]].border = b_border
    ws[row_range[7]].border = b_border

    wb.save('Final_Report.xlsx') #saves formatting


# function determines whether report is grow in or fall in

def process_button_command():
	#report_naming variable changes depending on which report is running below
	global report_naming, GIAS_report, GIBO_report, GIMO_report, FIMO_report, GIAS_report_sheet, GIBO_report_sheet, GIMO_report_sheet, FIMO_report_sheet
	if os.path.isfile('ViR-template.xlsx'):
		try:
			report_naming = "Grow-In_As-Surveyed"
			xml_conversion_script(GIAS_path)
			if os.path.isfile('Grow-In_As-Surveyed.xlsx'):
				feedback_text_1.configure(text = "Infringements Found" , bg = "OliveDrab4")
				GIAS_report = load_workbook(filename = 'Grow-In_As-Surveyed.xlsx')
				GIAS_report_sheet = GIAS_report.active
			else:
				feedback_text_1.configure(text = "No Infringements Found" , bg = "Orange")		
			if os.path.isfile('dataframe.xlsx'):					
				os.remove("dataframe.xlsx")
			
			report_naming = "Grow-In_Blow-Out"
			xml_conversion_script(GIBO_path)
			if os.path.isfile('Grow-In_Blow-Out.xlsx'):
				feedback_text_2.configure(text = "Infringements Found" , bg = "OliveDrab4")
				GIBO_report = load_workbook(filename = 'Grow-In_Blow-Out.xlsx')
				GIBO_report_sheet = GIBO_report.active
			else:
				feedback_text_2.configure(text = "No Infringements Found" , bg = "Orange")
			if os.path.isfile('dataframe.xlsx'):					
				os.remove("dataframe.xlsx")
			
			report_naming = "Grow-In_Max-Op"
			xml_conversion_script(GIMO_path)
			if os.path.isfile('Grow-In_Max-Op.xlsx'):
				feedback_text_3.configure(text = "Infringements Found" , bg = "OliveDrab4")
				GIMO_report = load_workbook(filename = 'Grow-In_Max-Op.xlsx')
				GIMO_report_sheet = GIMO_report.active
			else:
				feedback_text_3.configure(text = "No Infringements Found" , bg = "Orange")
			if os.path.isfile('dataframe.xlsx'):					
				os.remove("dataframe.xlsx")
			
			report_naming = "Fall-In-Max-Op"
			xml_conversion_script(FIMO_path)
			if os.path.isfile('Fall-In-Max-Op.xlsx'):
				feedback_text_4.configure(text = "Infringements Found" , bg = "OliveDrab4")
				FIMO_report = load_workbook(filename = 'Fall-In-Max-Op.xlsx')
				FIMO_report_sheet = FIMO_report.active
			else:
				feedback_text_4.configure(text = "No Infringements Found" , bg = "Orange")
			if os.path.isfile('dataframe.xlsx'):					
				os.remove("dataframe.xlsx")		
		
		except:
			errormsg( msg = "\nProblem loading XML")

		try:
			xlsx_collate()
			feedback_text_5.configure(text = "All Done! :)", fg = 'OliveDrab4')
		
		except:
			errormsg( msg = "\nError collating final spreadsheet")
			feedback_text_5.configure(text = "Error", fg = 'red')
	else:
		errormsg( msg = "\nCannot find 'ViR-Template.xlsx'.")
		feedback_text_5.configure(text = "Error", fg = 'red')

# pop up function

def errormsg(msg):
    popup = tk.Tk()
    popup.wm_title("!")
    label = ttk.Label(popup, text=msg, anchor="center")
    label.pack(side="top", fill="x", pady=10, padx=20)
    B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
    B1.pack(pady=20)



# -----------------------------------------GUI Tkinter App---------------------------------------------

# window creation

window = Tk()

window.title("ViR Reporting Tool")

window.geometry('440x250')

window.configure(bg = "grey25")


# Labels
GIAS_label = Label(window, text="Direct to 'Grow In As Surveyed' XML file:", fg = "ivory2")
GIAS_label.grid(column=0, row=0, pady=5)
GIAS_label.configure(bg = "grey25")

GIBO_label = Label(window, text="Direct to 'Grow In Blow Out' XML file:", fg = "ivory2")
GIBO_label.grid(column=0, row=2, pady=5)
GIBO_label.configure(bg = "grey25")

GIMO_label = Label(window, text="Direct to 'Grow In Max Op' XML file:", fg = "ivory2")
GIMO_label.grid(column=0, row=4, pady=5)
GIMO_label.configure(bg = "grey25")

FIMO_label = Label(window, text="Direct to 'Fall In Max Op' XML file:", fg = "ivory2")
FIMO_label.grid(column=0, row=6, pady=5)
FIMO_label.configure(bg = "grey25")


# Buttons
GIAS_button = tk.Button(window, text='Browse', command=GIAS_import, bg = "ivory4")
GIAS_button.grid(column=2, row=0, pady=5)

GIBO_button = tk.Button(window, text='Browse', command=GIBO_import, bg = "ivory4")
GIBO_button.grid(column=2, row=2, pady=5)

GIMO_button = tk.Button(window, text='Browse', command=GIMO_import, bg = "ivory4")
GIMO_button.grid(column=2, row=4, pady=5)

FIMO_button = tk.Button(window, text='Browse', command=FIMO_import, bg = "ivory4")
FIMO_button.grid(column=2, row=6, pady=5)

process_button = tk.Button(window, text='Process', height = 2, width = 10, 
				command = process_button_command, bg = "ivory4")
process_button.grid(row=7, column=1, rowspan=2, sticky=S, pady=20)


# feedback labels
feedback_text_1 = Label(window, width = 20, bg = "azure4")
feedback_text_1.grid(column = 1, row = 0, padx=(10, 10))

feedback_text_2 = Label(window, width = 20, bg = "azure4")
feedback_text_2.grid(column = 1, row = 2, padx=(10, 10))

feedback_text_3 = Label(window, width = 20, bg = "azure4")
feedback_text_3.grid(column = 1, row = 4, padx=(10, 10))

feedback_text_4 = Label(window, width = 20, bg = "azure4")
feedback_text_4.grid(column = 1, row = 6, padx=(10, 10))

feedback_text_5 = Label(window, width = 20, bg = "grey25", fg = "White")
feedback_text_5.grid(column = 0, row = 8, padx=(5, 5))

# opens the relevant sheets from the master template
try:
	template = load_workbook(filename = 'ViR-template.xlsx')
	GIAS_sheet = template['Grow-In (As-Surveyed MVCD)']
	GIBO_sheet = template['Blow-Out (NESC Blowout WSZ)']
	GIMO_sheet = template['Grow-In(CircuitThermal STE WSZ)']
	FIMO_sheet = template['Fall-In(CircuitThermal STE WSZ)']
except:
	errormsg( msg = "\nError! Could not find 'ViR-Template.xlsx'. \nPlease include this in the workspace and rerun the exe.")
	
window.mainloop()



