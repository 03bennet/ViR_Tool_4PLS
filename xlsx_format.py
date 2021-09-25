import time
import os
import ntpath
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Border, Alignment
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter import filedialog



#def xlsx_format(worksheet):

xlsx_file = 'Final_Report.xlsx'

# opening:
wb = load_workbook(filename = xlsx_file)

# center align column H in the default sheet:
ws = wb['Grow-In (As-Surveyed MVCD)']

mr = ws.max_row
mc = ws.max_column




for i in range(1, mr + 1):
    for j in range (1, mc + 1):
        c = ws.cell(row = i+9, column = j)
        print(c.value)

